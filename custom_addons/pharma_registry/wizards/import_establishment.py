# -*- coding: utf-8 -*-
"""Assistant d'import pour chaque type d'établissement pharmaceutique."""

import base64
import csv
import io

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - dépendance optionnelle
    load_workbook = None


class PharmaImportEstablishmentWizard(models.TransientModel):
    """Proposer un importeur générique pour officines, dépôts, etc."""

    _name = "pharma.import.establishment.wizard"
    _description = "Assistant d'import des établissements"

    type_etablissement = fields.Selection(
        [
            ("officine", "Officine"),
            ("depot", "Dépôt de médicaments"),
            ("grossiste", "Grossiste répartiteur"),
            ("agence", "Agence de promotion"),
            ("fabrication", "Établissement de fabrication"),
        ],
        string="Type d'établissement",
        required=True,
        default=lambda self: self.env.context.get("default_type_etablissement") or "officine",
    )
    file_data = fields.Binary(string="Fichier", required=True)
    filename = fields.Char(string="Nom du fichier")
    allow_update = fields.Boolean(string="Mettre à jour les fiches existantes", default=True)

    def action_import(self):
        """Créer ou mettre à jour des fiches à partir du fichier transmis."""

        self.ensure_one()
        rows = self._read_rows()
        if not rows:
            raise UserError(_("Le fichier ne contient aucune donnée."))

        model_name = {
            "officine": "pharma.officine",
            "depot": "pharma.depot",
            "grossiste": "pharma.grossiste",
            "agence": "pharma.agence",
            "fabrication": "pharma.fabrication",
        }[self.type_etablissement]

        Model = self.env[model_name]
        created = updated = 0

        for row in rows:
            vals = self._prepare_vals(row)
            if not vals.get("name"):
                raise UserError(_("La colonne 'Nom' est obligatoire pour chaque enregistrement."))

            domain = [("name", "=ilike", vals["name"])]
            record = Model.search(domain, limit=1)

            if record:
                if self.allow_update:
                    record.write(vals)
                    updated += 1
                continue
            Model.create(vals)
            created += 1

        message = _("%s créations, %s mises à jour") % (created, updated)
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import des établissements"),
                "message": message,
                "type": "success",
            },
        }

    # ------------------------------------------------------------------
    # Fonctions utilitaires
    # ------------------------------------------------------------------
    def _read_rows(self):
        """Convertir le fichier binaire en liste de dictionnaires."""

        decoded = base64.b64decode(self.file_data)
        if not self.filename:
            raise UserError(_("Veuillez fournir un nom de fichier."))
        extension = self.filename.lower().rsplit(".", 1)[-1]
        if extension == "csv":
            content = decoded.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            return [row for row in reader if any(row.values())]
        if extension in ("xlsx", "xlsm"):
            if load_workbook is None:
                raise UserError(_("Le module python openpyxl est requis pour lire les fichiers XLSX."))
            workbook = load_workbook(filename=io.BytesIO(decoded), read_only=True, data_only=True)
            sheet = workbook.active
            headers = []
            data_rows = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index == 0:
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    continue
                if not any(row):
                    continue
                record = {}
                for header, cell in zip(headers, row):
                    if not header:
                        continue
                    record[header] = cell if cell is not None else ""
                data_rows.append(record)
            return data_rows
        raise UserError(_("Format de fichier non supporté: %s" % extension))

    def _prepare_vals(self, row):
        """Mapper les valeurs brutes vers les champs du modèle, quel que soit le type."""

        base_vals = {}
        name_key = {
            "officine": "NOM DE L'OFFICINE",
            "depot": "NOM DU DEPOT",
            "grossiste": "NOM GROSSISTE REPARTITEUR",
            "agence": "NOM DE L'AGENCE DE PROMOTION",
            "fabrication": "NOM DE L'ETABLISSEMENT",
        }[self.type_etablissement]
        base_vals["name"] = self._get_value(row, name_key)

        base_vals["region_id"] = self._find_region(row.get("REGION"))
        base_vals["department_id"] = self._find_department(row.get("DEPARTEMENT"), base_vals["region_id"])
        base_vals["commune_id"] = self._find_commune(row.get("COMMUNE"), base_vals["department_id"])
        base_vals["quartier"] = self._required_value(row, "QUARTIER") or row.get("QUARTIER/VILLAGE/HAMEAU")
        base_vals["adresse"] = self._required_value(row, self._get_address_key())
        latitude, longitude = self._extract_coordinates(row)
        base_vals["latitude"] = latitude
        base_vals["longitude"] = longitude
        base_vals["observations"] = self._get_value(row, "OBSERVATIONS")

        handler = getattr(self, f"_prepare_{self.type_etablissement}")
        base_vals.update(handler(row))
        return base_vals

    # Traitements spécifiques -------------------------------------------------
    def _prepare_officine(self, row):
        """Extraire les colonnes propres aux officines."""

        return {
            "numero_telephone": self._required_value(row, "NUMERO TELEPHONE"),
            "annee_creation": self._required_int(row.get("ANNEE CREATION")),
            "annee_exploitation": self._required_int(
                row.get("ANNEE   D'EXPLOITATION") or row.get("ANNEE D'EXPLOITATION")
            ),
            "statut": self._map_statut(row.get("STATUT (TRANSFERT-RACHAT)")),
            "titulaire_nom": self._required_value(row, "PRENOM ET NOM TITULAIRE/PHARMACIEN RESPONSABLE"),
            "numero_ordre": self._required_value(row, "NUMERO D'INSCRIPTION ORDRE DES PHARMACIEN"),
            "sexe_titulaire": self._map_sexe(row.get("SEXE")),
            "tranche_age": self._map_tranche(row.get("TRANCHE D'AGE")),
            "nombre_assistants": self._required_int(row.get("NOMBRE D'ASSISTANTS")),
            "nombre_employe_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE PHARMACIEN")),
            "nombre_employe_non_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE NON PHARMACIEN")),
            "nombre_agent_securite": self._required_int(row.get("NOMBRE D'AGENT DE SECURITE")),
            "nombre_agent_hygiene": self._required_int(row.get("NOMBRE D'AGENT D'HYGIENE")),
            "chiffre_affaire": self._to_float(row.get("CHIFFRE D'AFFAIRE")),
            "nombre_vehicule": self._to_int(row.get("NOMBRE DE VEHICULE(LIVRAISON ET TRANSFERT)")),
        }

    def _prepare_depot(self, row):
        """Extraire les colonnes propres aux dépôts."""

        return {
            "numero_telephone": self._required_value(row, "NUMERO TELEPHONE"),
            "annee_ouverture": self._required_int(row.get("ANNEE D'OUVERTURE")),
            "responsable_nom": self._required_value(row, "PRENOM ET NOM RESPONSABLE/DEPOSITAIRE"),
            "sexe_responsable": self._map_sexe(row.get("SEXE")),
        }

    def _prepare_grossiste(self, row):
        """Extraire les colonnes propres aux grossistes répartiteurs."""

        return {
            "numero_telephone": self._required_value(row, "NUMERO TELEPHONE"),
            "annee_ouverture": self._required_int(row.get("ANNEE D'OUVERTURE")),
            "responsable_nom": self._required_value(row, "PRENOM ET NOM RESPONSABLE/DIRECTEUR"),
            "nombre_employe_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE PHARMACIEN")),
            "nombre_employe_non_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE NON PHARMACIEN")),
            "nombre_agent_securite": self._required_int(row.get("NOMBRE D'AGENT DE SECURITE")),
            "nombre_agent_hygiene": self._required_int(row.get("NOMBRE D'AGENT D'HYGIENE")),
            "chiffre_affaire": self._to_float(row.get("CHIFFRE D'AFFAIRE")),
            "nombre_vehicule": self._to_int(row.get("NOMBRE DE VEHICULE(LIVRAISON ET TRANSFERT)")),
        }

    def _prepare_agence(self, row):
        """Extraire les colonnes propres aux agences de promotion."""

        return {
            "numero_telephone": self._required_value(row, "NUMERO TELEPHONE"),
            "annee_ouverture": self._required_int(row.get("ANNEE D'OUVERTURE")),
            "numero_agrement": self._required_value(row, "NUMERO DE L'AGREMENT"),
            "date_agrement": self._to_date(row.get("DATE DE L'AGREMENT"), required=True),
            "pharmacien_responsable": self._required_value(row, "PRENOM ET NOM  DU PHARMACIEN RESPONSABLE"),
            "nombre_employe_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE PHARMACIEN")),
            "nombre_employe_non_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE NON PHARMACIEN")),
            "chiffre_affaire": self._to_float(row.get("CHIFFRE D'AFFAIRE")),
            "laboratoire_represente": self._get_value(row, "NOM DU LABORATOIRE REPRESENTE"),
        }

    def _prepare_fabrication(self, row):
        """Extraire les colonnes propres aux sites de fabrication."""

        return {
            "numero_telephone": self._required_value(row, "NUMERO TELEPHONE"),
            "annee_ouverture": self._required_int(row.get("ANNEE D'OUVERTURE")),
            "responsable_nom": self._required_value(row, "PRENOM ET NOM RESPONSABLE"),
            "nombre_employe_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE PHARMACIEN")),
            "nombre_employe_non_pharmacien": self._required_int(row.get("NOMBRE D'EMPLOYE NON PHARMACIEN")),
            "nombre_agent_securite": self._required_int(row.get("NOMBRE D'AGENT DE SECURITE")),
            "nombre_agent_hygiene": self._required_int(row.get("NOMBRE D'AGENT D'HYGIENE")),
            "chiffre_affaire": self._to_float(row.get("CHIFFRE D'AFFAIRE")),
        }

    def _extract_coordinates(self, row):
        """Retourner un tuple (latitude, longitude) selon les colonnes disponibles."""

        latitude = self._to_coordinate(row.get("LATITUDE"))
        longitude = self._to_coordinate(row.get("LONGITUDE"))

        if latitude is False or longitude is False:
            raw = self._get_value(row, "POINTS DE GEOLOCALISATION")
            latitude, longitude = self._parse_points(raw)

        return latitude, longitude

    # Méthodes utilitaires --------------------------------------------------
    def _find_region(self, name):
        if not name:
            return False
        region = self.env["pharma.geo.region"].search([("name", "=ilike", name.strip())], limit=1)
        if not region:
            raise UserError(_("La région '%s' est introuvable. Veuillez l'importer au préalable.") % name)
        return region.id

    def _find_department(self, name, region_id):
        if not name:
            return False
        domain = [("name", "=ilike", name.strip())]
        if region_id:
            domain.append(("region_id", "=", region_id))
        department = self.env["pharma.geo.department"].search(domain, limit=1)
        if not department:
            raise UserError(_("Le département '%s' est introuvable pour la région sélectionnée.") % name)
        return department.id

    def _find_commune(self, name, department_id):
        if not name:
            return False
        if not department_id:
            raise UserError(_("Impossible d'associer la commune '%s' sans département." % name))
        commune = self.env["pharma.geo.commune"].search(
            [
                ("name", "=ilike", name.strip()),
                ("department_id", "=", department_id),
            ],
            limit=1,
        )
        if not commune:
            raise UserError(_("La commune '%s' est introuvable pour le département fourni.") % name)
        return commune.id

    def _get_value(self, row, key):
        """Retourner la valeur texte nettoyée lorsqu'elle existe."""

        value = row.get(key)
        if isinstance(value, str):
            return value.strip()
        return value

    def _required_value(self, row, key):
        value = self._get_value(row, key)
        if not value:
            raise UserError(_("La colonne '%s' est obligatoire." % key))
        return value

    def _get_adresse(self, row):
        """Identifier la colonne d'adresse en fonction du type sélectionné."""

        mapping = {
            "officine": "ADRESSE  EXACTE DE L'OFFICINE",
            "depot": "ADRESSE  EXACTE DU DEPOT",
            "grossiste": "ADRESSE  EXACTE",
            "agence": "ADRESSE  EXACTE",
            "fabrication": "ADRESSE  EXACTE",
        }
        key = mapping[self.type_etablissement]
        return self._get_value(row, key)

    def _map_statut(self, value):
        """Normaliser le statut texte pour le ramener aux clés de sélection."""

        if not value:
            return False
        value = str(value).strip().lower()
        if "trans" in value:
            return "transfert"
        if "rach" in value:
            return "rachat"
        return "autre"

    def _map_sexe(self, value):
        """Normaliser les valeurs de sexe afin de correspondre à l'énumération."""

        if not value:
            return "na"
        value = str(value).strip().lower()
        if value.startswith("f"):
            return "f"
        if value.startswith("m"):
            return "m"
        return "na"

    def _map_tranche(self, value):
        """Convertir les tranches d'âge textuelles en codes internes."""

        if not value:
            return "na"
        value = str(value).strip().lower()
        mapping = {
            "18": "moins_30",
            "25": "moins_30",
            "30": "30_39",
            "35": "30_39",
            "40": "40_49",
            "45": "40_49",
            "50": "50_59",
            "55": "50_59",
            "60": "60_plus",
        }
        for key, code in mapping.items():
            if key in value:
                return code
        if "60" in value or "plus" in value:
            return "60_plus"
        return "na"

    def _to_int(self, value):
        """Convertir en entier de manière sûre lorsque c'est possible."""

        try:
            return int(float(value)) if value not in (None, "") else False
        except (TypeError, ValueError):
            return False

    def _to_float(self, value):
        """Convertir en nombre décimal de manière sûre lorsque c'est possible."""

        try:
            return float(value) if value not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _to_date(self, value):
        """Analyser les formats de date courants et retourner un objet date."""

        if not value:
            return False
        if hasattr(value, "isoformat"):
            return value
        try:
            from datetime import datetime

            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(str(value), fmt).date()
                except ValueError:
                    continue
        except Exception:  # pragma: no cover - solution de repli
            return False
        raise UserError(_("Format de date inconnu: %s" % value))

    def _to_coordinate(self, value):
        """Convertir une valeur en flottant, retourner False si vide ou invalide."""

        if value in (None, ""):
            return False
        try:
            return float(value)
        except (TypeError, ValueError):
            return False

    def _parse_points(self, raw_points):
        """Décoder une chaîne "lon,lat" en coordonnées numériques."""

        if not raw_points:
            return False, False
        parts = [chunk.strip() for chunk in str(raw_points).split(",") if chunk.strip()]
        if len(parts) < 2:
            return False, False
        try:
            longitude = float(parts[0])
            latitude = float(parts[1])
        except ValueError:
            return False, False
        return latitude, longitude
