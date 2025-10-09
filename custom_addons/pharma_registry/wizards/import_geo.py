# -*- coding: utf-8 -*-
import base64
import csv
import io
import unicodedata

from odoo import models, fields, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - dépendance optionnelle
    load_workbook = None


class PharmaImportGeoWizard(models.TransientModel):
    """Permettre aux administrateurs de charger en masse régions, départements et communes."""

    _name = "pharma.import.geo.wizard"
    _description = "Assistant d'import géographique"

    file_data = fields.Binary(string="Fichier", required=True)
    filename = fields.Char(string="Nom du fichier")

    def action_import(self):
        """Parcourir le fichier reçu et créer ou compléter la hiérarchie géographique."""

        self.ensure_one()
        rows = self._read_rows()
        if not rows:
            raise UserError(_("Le fichier ne contient aucune donnée."))

        Region = self.env["pharma.geo.region"]
        Department = self.env["pharma.geo.department"]
        Commune = self.env["pharma.geo.commune"]

        created_regions = created_departments = created_communes = 0

        for row in rows:
            region_name = self._get_first_value(row, "REGION", "Région")
            department_name = self._get_first_value(row, "DEPARTEMENT", "Département", "Department")
            commune_name = self._get_first_value(row, "COMMUNE", "Commune")

            if not region_name:
                raise UserError(_("La colonne REGION est obligatoire."))

            region = Region.search([("name", "=ilike", region_name)], limit=1)
            if not region:
                region = Region.create({"name": region_name})
                created_regions += 1

            department = False
            if department_name:
                department = Department.search(
                    [
                        ("name", "=ilike", department_name),
                        ("region_id", "=", region.id),
                    ],
                    limit=1,
                )
                if not department:
                    department = Department.create(
                        {
                            "name": department_name,
                            "region_id": region.id,
                        }
                    )
                    created_departments += 1

            if commune_name and department:
                commune = Commune.search(
                    [
                        ("name", "=ilike", commune_name),
                        ("department_id", "=", department.id),
                    ],
                    limit=1,
                )
                if not commune:
                    Commune.create(
                        {
                            "name": commune_name,
                            "department_id": department.id,
                            "region_id": region.id,
                        }
                    )
                    created_communes += 1

        message = _(
            "%s régions créées, %s départements créés, %s communes créées."
        ) % (created_regions, created_departments, created_communes)

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import géographique"),
                "message": message,
                "type": "success",
            },
        }

    def _read_rows(self):
        """Charger les lignes CSV/XLSX et retourner des dictionnaires normalisés."""

        decoded = base64.b64decode(self.file_data)
        if not self.filename:
            raise UserError(_("Veuillez fournir un nom de fichier."))
        extension = self.filename.lower().rsplit(".", 1)[-1]
        if extension == "csv":
            content = decoded.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            return [self._normalize_row(row) for row in reader if any(row.values())]
        if extension in ("xlsx", "xlsm"):
            if load_workbook is None:
                raise UserError(_("Le module python openpyxl est requis pour lire les fichiers XLSX."))
            workbook = load_workbook(filename=io.BytesIO(decoded), read_only=True, data_only=True)
            sheet = workbook.active
            headers = []
            data_rows = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index == 0:
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    continue
                if not any(row):
                    continue
                record = {}
                for header, cell in zip(headers, row):
                    if not header:
                        continue
                    record[header] = cell if cell is not None else ""
                data_rows.append(self._normalize_row(record))
            return data_rows
        raise UserError(_("Format de fichier non supporté: %s" % extension))

    # ------------------------------------------------------------------
    # Fonctions utilitaires
    # ------------------------------------------------------------------
    def _normalize_row(self, row):
        """Copier la ligne et ajouter des variantes d'en-têtes sans accents."""

        normalized = dict(row)
        for key, value in list(row.items()):
            normalized_key = self._normalize_header(key)
            if normalized_key and normalized_key not in normalized:
                normalized[normalized_key] = value
        return normalized

    @staticmethod
    def _normalize_header(header):
        """Transformer un en-tête en identifiant ASCII majuscule."""

        if not header:
            return ""
        normalized = unicodedata.normalize("NFKD", header)
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = normalized.replace(" ", "_")
        return normalized.upper()

    def _get_first_value(self, row, *keys):
        """Récupérer la première valeur non vide parmi les alias d'en-tête."""

        for key in keys:
            if not key:
                continue
            variants = {
                key,
                key.lower(),
                key.upper(),
                key.title(),
                self._normalize_header(key),
            }
            for variant in variants:
                value = row.get(variant)
                if value not in (None, ""):
                    return str(value).strip()
        return ""
